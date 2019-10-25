from openpyxl import Workbook,load_workbook
from PIL import Image, ImageFilter

""" 图像信息
img=Image.open('Learning-Python-100-Days/day15/images/person.jpg')
print(img.format)
print(img.size)
print(img.mode)
img.show()
"""
""" 裁剪图像
img=Image.open('Learning-Python-100-Days/day15/images/person.jpg')
img.crop((80,20,310,360)).show()
"""

""" 生成缩略图
img=Image.open('Learning-Python-100-Days/day15/images/person.jpg')
# 缩略图保持比例，取决于传入的宽高中较小的一边
img.thumbnail((128,128))
img.show()
"""

""" 粘贴图像
img1=Image.open('Learning-Python-100-Days/day15/images/person.jpg')
img2=Image.open('Learning-Python-100-Days/day15/images/person2.jpg')
img1_head=img1.crop((120,130,310,260))
width,height=img1_head.size[0],img1_head.size[1]
print(img1_head.size)
print(img2.size)
img2.paste(img1_head,(0,0,width,height))
img2.show()
"""

""" 旋转/反转图像
img=Image.open('Learning-Python-100-Days/day15/images/person.jpg')
img.rotate(90).show() #逆时针旋转90度
img.transpose(Image.FLIP_LEFT_RIGHT).show() #左右翻转
img.transpose(Image.FLIP_TOP_BOTTOM).show() #上下翻转
"""

""" 操作像素
img=Image.open('Learning-Python-100-Days/day15/images/person.jpg')
for x in range(10,100):
    for y in range(20,200):# 可以看作从（10，20）坐标到（100，200）坐标画一个矩形
        img.putpixel((x,y),(255,255,255))
img.show()
"""

""" 滤镜
img=Image.open('Learning-Python-100-Days/day15/images/person.jpg')
img.filter(ImageFilter.CONTOUR).show()
"""

# %%
""" xls文件的创建以及读写
# ABCD是列，1234是行
wb = Workbook()  # 新建一个工作簿

ws = wb.create_sheet("Mysheet")  # 默认末尾插入 再带一个参数时：0插入第一个，-1倒数第二个

ws.title = "MyShit"  # 修改表名
ws.sheet_properties.tabColor = "1072BA"  # 修改颜色
print(wb.sheetnames)

#------------------
ws['A4'].value = 'saseds'  # 第一种方法：直接指定单元格赋值
a = ws['A4'].value  # 指定单元格取值

b=ws.cell(row=4,column=2,value=10).value #第二种方式：指定行和列
#------------------

#------------------
cell_range=ws['A1':'C2'] #取范围内的单元
b2=cell_range[1][1].value #理解一下此处的位置关系

cell_all=tuple(ws.rows) #取所有的单元格

for row in ws.values:  #取所有单元格中的值
   for value in row:
     print(value) 
#------------------
print(cell_all)
wb.save('Learning-Python-100-Days\\day15\\oxo.xlsx') # 会覆盖现有文件且不提醒

wb2=load_workbook('Learning-Python-100-Days\\day15\\oxo.xlsx') #加载工作簿

print(wb2.sheetnames)
"""

